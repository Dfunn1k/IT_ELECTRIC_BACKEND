from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.shortcuts import render
from django.utils import timezone
import time
import pandas as pd
from openpyxl import load_workbook
import numpy as np

from concurrent.futures import ThreadPoolExecutor

from rest_framework import generics, status
from rest_framework.authtoken.models import Token
from rest_framework.authtoken.serializers import AuthTokenSerializer
from rest_framework.generics import CreateAPIView, ListAPIView
from rest_framework.parsers import FileUploadParser, FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import MeasurementER, MeasurementTB, Engine, ElectricalResult, TestER, TestTB, TransientBoot, AverageMeasurement
from .serializers import MeasurementERSerializer, MeasurementTBSerializer, EngineSerializer, ElectricalResultSerializer, TestERSerializer, TestTBSerializer, TransientBootSerializer, UserSerializer

from datetime import datetime

import multiprocessing
import concurrent.futures
from .utils import create_objects, process_array, read_excel_chunk, data_avarage


class GetUserList(ListAPIView):
    permission_classes = [IsAuthenticated]
    queryset = User.objects.all()
    serializer_class = UserSerializer


class LoginView(APIView):
    def post(self, request):
        # Verifica las credenciales del usuario
        serializer = AuthTokenSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']

        # Crea o recupera el token del usuario
        token, created = Token.objects.get_or_create(user=user)

        # Devuelve el token y el ID del usuario en la respuesta
        return Response({'token': token.key, 'user_id': user.id})


class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        # Elimina el token del usuario
        Token.objects.filter(user=request.user).delete()
        return Response({'detail': 'El token fue eliminado'}, status=204)


class CreateEngineView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, format=None):
        # Utilizar el JSONParser para procesar los datos en formato JSON
        parser = JSONParser()
        data = parser.parse(request)

        serializer = EngineSerializer(data=data)
        if serializer.is_valid():
            # Obtener o crear el motor según su número nombre
            try:
                engine = Engine.objects.get(name=data["name"],
                                            user=request.user)
                response_data = EngineSerializer(engine).data
            except (Engine.DoesNotExist, KeyError):
                # motor = serializer.save(usuario=request.user)
                response_data = serializer.save(user=request.user)

            return Response(response_data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class GetEnginesUserView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, user_pk):
        engines = Engine.objects.filter(user__pk=user_pk)
        serializer = EngineSerializer(engines, many=True)
        return Response(serializer.data)


class GetMeasurementsERView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, test_er):
        """Esta función obtiene los datos de las columnas mostradas, re"""
        measurements = MeasurementER.objects.filter(
            test_electrical_result_fk=test_er)
        data = {
            # 'item': list(mediciones.values_list('item', flat=True)),
            'time': list(measurements.values_list('time', flat=True)),
            'mag_v1': list(map(lambda x: round(x, 2), measurements.values_list('mag_v1', flat=True))),
            'mag_v2': list(map(lambda x: round(x, 2), measurements.values_list('mag_v2', flat=True))),
            'mag_v3': list(map(lambda x: round(x, 2), measurements.values_list('mag_v3', flat=True))),
            'mag_i1': list(map(lambda x: round(x, 2), measurements.values_list('mag_i1', flat=True))),
            'mag_i2': list(map(lambda x: round(x, 2), measurements.values_list('mag_i2', flat=True))),
            'mag_i3': list(map(lambda x: round(x, 2), measurements.values_list('mag_i3', flat=True)))
        }
        return Response(data)


class GetMeasurementsTBView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, test_tb):
        measurements = MeasurementTB.objects.filter(
            test_transient_boot_fk=test_tb)
        data = {
            'time': list(map(lambda x: round(x, 2), measurements.values_list('time', flat=True))),
            'ia': list(map(lambda x: round(x, 2), measurements.values_list('ia', flat=True))),
            'ib': list(map(lambda x: round(x, 2), measurements.values_list('ib', flat=True))),
            'ic': list(map(lambda x: round(x, 2), measurements.values_list('ic', flat=True))),
            'va': list(map(lambda x: round(x, 2), measurements.values_list('va', flat=True))),
            'vb': list(map(lambda x: round(x, 2), measurements.values_list('vb', flat=True))),
            'vc': list(map(lambda x: round(x, 2), measurements.values_list('vc', flat=True)))
        }
        return Response(data)


class UploadMeasurementsERView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, pk):
        # Verificamos que se este enviando un archivo en la petición
        if "file" not in request.data:
            return Response(
                {"error": "No file R.E. provided"},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.data["file"]

        # Verificamos que el archivo tenga una extensión xlsx
        if not file.name.endswith('.xlsx'):
            return Response(
                {"error": "El archivo debe ser un archivo de Excel (.xlsx)"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener la data necesaria para crear el objeto testRE
        electrical_result_fk = pk
        obj_electrical_result = ElectricalResult.objects.get(
            pk=electrical_result_fk)
        # Leer el archivo de Excel y obtener las mediciones
        wb = load_workbook(file, read_only=True)
        # columnas_deseadas = [2,3,4]

        sheet_name = wb.sheetnames[0].rsplit("-", 2)
        sheet_name = ":".join(sheet_name)
        test_date_time = datetime.strptime(sheet_name, "%Y-%m-%dT%H:%M:%S")
        test_date_time = timezone.make_aware(test_date_time)
        try:
            test_electrical_result = TestER.objects.create(
                electrical_result_fk=obj_electrical_result,
                test_date_time=test_date_time
            )
        except ValidationError as e:
            test_electrical_result = TestER.objects.get(
                electrical_result_fk=obj_electrical_result, test_date_time=test_date_time)
            test_electrical_result.delete()
            return Response({'error': f"{e}, tambien se elimino el testER", }, status=400)
        df = pd.read_excel(file)
        number_cores = multiprocessing.cpu_count()

        try:
            array = df.to_numpy()
            promedio = data_avarage(array, electrical_result_fk)
            obj_average = AverageMeasurement(test_electrical_result_fk=test_electrical_result,
                                             ab=promedio["voltage"]["ab"],
                                             bc=promedio["voltage"]["bc"],
                                             ca=promedio["voltage"]["ca"],
                                             avg=promedio["voltage"]["avg"],
                                             value=promedio["voltage"]["value"],
                                             unbalance=promedio["unbalance"],
                                             thdv_a=promedio["distorsion"]["thdv_a"],
                                             thdv_b=promedio["distorsion"]["thdv_b"],
                                             thdv_c=promedio["distorsion"]["thdv_c"],
                                             thdv_avg=promedio["distorsion"]["thdv_avg"],
                                             thdi_a=promedio["distorsion"]["thdi_a"],
                                             thdi_b=promedio["distorsion"]["thdi_b"],
                                             thdi_c=promedio["distorsion"]["thdi_c"],
                                             thdi_avg=promedio["distorsion"]["thdi_avg"],
                                             tdv_a=promedio["full_distorsion"]["tdv_a"],
                                             tdv_b=promedio["full_distorsion"]["tdv_b"],
                                             tdv_c=promedio["full_distorsion"]["tdv_c"],
                                             tdv_avg=promedio["full_distorsion"]["tdv_avg"],
                                             tdi_a=promedio["full_distorsion"]["tdi_a"],
                                             tdi_b=promedio["full_distorsion"]["tdi_b"],
                                             tdi_c=promedio["full_distorsion"]["tdi_c"],
                                             tdi_avg=promedio["full_distorsion"]["tdi_avg"],
                                             current_a=promedio["current_level"]["current_a"],
                                             current_b=promedio["current_level"]["current_b"],
                                             current_c=promedio["current_level"]["current_c"],
                                             current_avg=promedio["current_level"]["current_avg"],
                                             current_nominal=promedio["current_level"]["current_nominal"],
                                             current_unbalance=promedio["current_unbalance"],
                                             load_percen_avg=promedio["efficiency"]["load_percen_avg"],
                                             lsskw_avg=promedio["efficiency"]["lsskw_avg"],
                                             eff_avg=promedio["efficiency"]["eff_avg"],
                                             sideband_amplitud_db=promedio["spectrum"]["sideband_amplitud_db"],
                                             sideband_freq_hz=promedio["spectrum"]["sideband_freq_hz"],
                                             vab_fase=promedio["symetrical_components"]["vab_fase"],
                                             vbc_fase=promedio["symetrical_components"]["vbc_fase"],
                                             vca_fase=promedio["symetrical_components"]["vca_fase"],
                                             unbalance_voltage=promedio["symetrical_components"]["unbalance_voltage"],
                                             va1_amplitud=promedio["symetrical_components"]["va1_amplitud"],
                                             va2_amplitud=promedio["symetrical_components"]["va2_amplitud"],
                                             va1_fase=promedio["symetrical_components"]["va1_fase"],
                                             va2_fase=promedio["symetrical_components"]["va2_fase"],
                                             ia_fase=promedio["symetrical_components"]["ia_fase"],
                                             ib_fase=promedio["symetrical_components"]["ib_fase"],
                                             ic_fase=promedio["symetrical_components"]["ic_fase"],
                                             unbalance_current=promedio["symetrical_components"]["unbalance_current"],
                                             ia1_amplitud=promedio["symetrical_components"]["ia1_amplitud"],
                                             ia2_amplitud=promedio["symetrical_components"]["ia2_amplitud"],
                                             ia1_fase=promedio["symetrical_components"]["ia1_fase"],
                                             ia2_fase=promedio["symetrical_components"]["ia2_fase"])
            obj_average.save()
        except:
            test_electrical_result.delete()
            print("Hubo un error al crear el objeto Measurements")

        split_array = np.array_split(array, number_cores)
        data_with_test = [(split_array[i], test_electrical_result)
                          for i in range(number_cores)]
        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=number_cores) as executor:
                executor.map(process_array, data_with_test)
        except:
            test_electrical_result.delete()
            print("hubo un error al leer el excel")

        return Response(
            {"message": "Las mediciones E.R han sido creadas exitosamente"},
            status=status.HTTP_201_CREATED,
        )


class UploadMeasurementsTBView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, pk):
        if "file" not in request.data:
            return Response(
                {"error": "No file T.A. provided"},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.data["file"]

        # verificamos que el archivo tenga una extensión xlsx
        if not file.name.endswith('.xlsx'):
            return Response(
                {"error": "El archivo debe ser un archivo de Excel (.xlsx)"},
                status=status.HTTP_400_BAD_REQUEST
            )

        transient_boot_fk = pk
        obj_transient_boot = TransientBoot.objects.get(
            pk=transient_boot_fk)
        wb = load_workbook(file, read_only=True)
        print(wb.sheetnames)
        sheet_name = wb.sheetnames[0].rsplit("-", 2)

        sheet_name = ":".join(sheet_name)
        test_date_time = datetime.strptime(sheet_name, "%Y-%m-%dT%H:%M:%S")
        test_date_time = timezone.make_aware(test_date_time)

        try:
            test_transient_boot = TestTB.objects.create(
                transient_boot_fk=obj_transient_boot,
                test_date_time=test_date_time
            )
            print("llegue a entrar al try")
        except ValidationError as e:
            test_transient_boot = TestTB.objects.get(
                transient_boot_fk=obj_transient_boot, test_date_time=test_date_time)
            test_transient_boot.delete()
            print("entre al except")
            return Response({'error': f"{e}, tambien se elimino el testTB"}, status=400)
        print("Sali del except")
        with ThreadPoolExecutor() as executor:
            futures = []
            for chunk in read_excel_chunk(wb, 1000, test_transient_boot):
                futures.append(executor.submit(create_objects, chunk))
            for future in futures:
                future.result()

        return Response(
            {"message": "Las mediciones T.B. han sido creadas exitosamente"},
            status=status.HTTP_201_CREATED,
        )


class DeleteEngineView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            engine = Engine.objects.get(pk=pk)
            engine.delete()
            return Response({"message": f"El motor con el id '{pk}' fue correctamente eliminado"}, status=status.HTTP_204_NO_CONTENT)
        except Engine.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)


class EditEngineView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        engine = Engine.objects.get(pk=pk)
        serializer = EngineSerializer(engine, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        engine = Engine.objects.get(pk=pk)
        serializer = EngineSerializer(engine, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DeleteTestERView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            test_electrical_result = TestER.objects.get(pk=pk)
            test_electrical_result.delete()
            return Response({"message": f"Las mediciones del test '{pk}' fueron correctamente eliminadas"}, status=status.HTTP_204_NO_CONTENT)
        except TestER.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)


class DeleteTestTBView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            test_transient_boot = TestTB.objects.get(pk=pk)
            test_transient_boot.delete()
            return Response({"message": f"Las mediciones del test '{pk}' fueron correctamente eliminadas"}, status=status.HTTP_204_NO_CONTENT)
        except TestTB.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)


class AverageView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, test_er):
        test_er_obj = TestER.objects.get(pk=test_er)
        electrical_result_obj = test_er_obj.electrical_result_fk
        engine_pk = electrical_result_obj.engine_fk.pk
        promedio = AverageMeasurement.objects.get(
            test_electrical_result_fk=test_er)
        tests = TestER.objects.filter(
            electrical_result_fk__engine_fk=engine_pk)
        history = {}
        for test in tests:
            value = AverageMeasurement.objects.get(
                test_electrical_result_fk=test.pk)
            history[test.pk] = {
                'value': round(value.value, 2),
                'fecha': test.test_date_time.strftime("%d/%m/%Y"),
                'hora': test.test_date_time.strftime("%H:%M:%S"),
            }

        gauge = {
            "minValue": 0,
            "maxValue": 120,
            "value": round(promedio.value, 2)
        }

        data = {
            "popupVoltage": {
                "values": {
                    "ab": round(promedio.ab, 2),
                    "bc": round(promedio.bc, 2),
                    "ca": round(promedio.ca, 2),
                    "avg": round(promedio.avg, 2),
                    "rated": electrical_result_obj.engine_fk.voltage_rating,
                    "value": round(promedio.value, 2)
                },
                "gauge": gauge,
                "history": history
            },
            "popupUnbalance": {
                "values": {
                    "ab": round(promedio.ab, 2),
                    "bc": round(promedio.bc, 2),
                    "ca": round(promedio.ca, 2),
                    "avg": round(promedio.avg, 2),
                    "unbalance": round(promedio.unbalance, 2),
                    "nemaDerating": electrical_result_obj.engine_fk.amps_rating,
                },
                "gauge": gauge,
                "history": history
            },
            "popupDistorsion": {
                "tdhv": {
                    "a": round(promedio.thdv_a, 2),
                    "b": round(promedio.thdv_b, 2),
                    "c": round(promedio.thdv_c, 2),
                    "avg": round(promedio.thdv_avg, 2),
                },
                "tdhi": {
                    "a": round(promedio.thdi_a, 2),
                    "b": round(promedio.thdi_b, 2),
                    "c": round(promedio.thdi_c, 2),
                    "avg": round(promedio.thdi_avg, 2)
                },
                "history": history
            },
            "popupFullDistorsion": {
                "tdv": {
                    "a": round(promedio.tdv_a, 2),
                    "b": round(promedio.tdv_b, 2),
                    "c": round(promedio.tdv_c, 2),
                    "avg": round(promedio.tdv_avg, 2)
                },
                "tdi": {
                    "a": round(promedio.tdi_a, 2),
                    "b": round(promedio.tdi_b, 2),
                    "c": round(promedio.tdi_c, 2),
                    "avg": round(promedio.tdi_avg, 2)
                },
                "gauge": gauge,
                "history": history
            },
            "popupCurrentLevel": {
                "values": {
                    "a": round(promedio.current_a, 2),
                    "b": round(promedio.current_b, 2),
                    "c": round(promedio.current_c, 2),
                    "avg": round(promedio.current_avg, 2),
                    "nominalCurrent": round(promedio.current_nominal, 2),
                    "rated": electrical_result_obj.engine_fk.voltage_rating
                },
                "gauge": gauge,
                "history": history
            },
            "popupCurrentUnbalance": {
                "values": {
                    "a": round(promedio.current_a, 2),
                    "b": round(promedio.current_b, 2),
                    "c": round(promedio.current_c, 2),
                    "avg": round(promedio.current_avg, 2),
                    "currentUnbalance": round(promedio.current_unbalance, 2),
                    "rated": electrical_result_obj.engine_fk.voltage_rating
                },
                "gauge": gauge,
                "history": history
            },
            "popupEfficiency": {
                "values": {
                    "load": round(promedio.load_percen_avg, 2),
                    "losses": round(promedio.lsskw_avg, 2),
                    "efficiency": round(promedio.eff_avg, 2)
                },
                "gauge": gauge,
                "history": history
            },
            "popupLoad": {
                "values": {
                    "efficiency": round(promedio.eff_avg, 2),
                    "speed": electrical_result_obj.engine_fk.speed_rpm,
                    "nemaDerating": electrical_result_obj.engine_fk.amps_rating,
                    "load": round(promedio.load_percen_avg, 2)
                },
                "gauge": gauge,
                "history": history
            },
            "popupSpectrum": {
                "values": {
                    "sideAmplitude": round(promedio.sideband_amplitud_db, 2),
                    "sidebandFreq": round(promedio.sideband_freq_hz, 2),
                    "fundFreq": electrical_result_obj.engine_fk.freq_hz,
                },
                "history": history
            },
            "SymmetricalComponents": [
                {
                    "section_left": [
                        {
                            "title": 'Vab',
                            "values": {
                                "amplitud": f"{round(promedio.ab,2)} [V]",
                                "fase": f"{round(promedio.vab_fase,2)}°"
                            }
                        },
                        {
                            "title": 'Vbc',
                            "values": {
                                "amplitud": f"{round(promedio.bc,2)} [V]",
                                "fase": f"{round(promedio.vbc_fase,2)}°"
                            }
                        },
                        {
                            "title": "Vca",
                            "values": {
                                "amplitud": f"{round(promedio.ca,2)} [V]",
                                "fase": f"{round(promedio.vca_fase,2)}°"
                            }
                        }],
                    "section_mid": round(promedio.unbalance_voltage, 2),
                    "section_right": [
                        {
                            "title": "VA1",
                            "values": {
                                "amplitud": round(promedio.va1_amplitud, 2),
                                "fase": round(promedio.va1_fase, 2)
                            }
                        },
                        {
                            "title": "VA2",
                            "values": {
                                "amplitud": round(promedio.va2_amplitud, 2),
                                "fase": round(promedio.va2_fase, 2)
                            }
                        }
                    ]
                },
                {
                    "section_left": [
                        {
                            "title": "IA",
                            "values": {
                                "amplitud": f"{round(promedio.current_a, 2)} [I]",
                                "fase": f"{round(promedio.ia_fase,2)}°"
                            }
                        },
                        {
                            "title": "IB",
                            "values": {
                                "amplitud": f"{round(promedio.current_b,2)} [I]",
                                "fase": f"{round(promedio.ib_fase, 2)}°"
                            }
                        },
                        {
                            "title": "IC",
                            "values": {
                                "amplitud": f"{round(promedio.current_c,2)} [I]",
                                "fase": f"{round(promedio.ic_fase,2)}°"
                            }
                        }
                    ],
                    "section_mid": round(promedio.unbalance_current, 2),
                    "section_right": [
                        {
                            "title": "IA1",
                            "values": {
                                "amplitud": round(promedio.ia1_amplitud, 2),
                                "fase": round(promedio.ia1_fase, 2)
                            }
                        },
                        {
                            "title": "IA2",
                            "values": {
                                "amplitud": round(promedio.ia2_amplitud, 2),
                                "fase": round(promedio.ia2_fase, 2)
                            }
                        }
                    ]
                }
            ]
        }

        return Response(data)


class MainView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, test_er):
        test_er_obj = TestER.objects.get(pk=test_er)
        electrical_result_obj = test_er_obj.electrical_result_fk
        engine_pk = electrical_result_obj.engine_fk.pk
        promedio = AverageMeasurement.objects.get(
            test_electrical_result_fk=test_er)
        tests = TestER.objects.filter(
            electrical_result_fk__engine_fk=engine_pk)

        data = {
            "engine": {
                "name": electrical_result_obj.engine_fk.name,
                "voltage": f"{electrical_result_obj.engine_fk.voltage_rating}",
                "current": f"{electrical_result_obj.engine_fk.amps_rating}",
                "kW": f"{electrical_result_obj.engine_fk.power_out_kw}",
                "RPM": f"{electrical_result_obj.engine_fk.speed_rpm}",
            },
            "Voltage": round(promedio.avg, 2),
            "Current": round(promedio.current_avg, 2),
            "FP": 0.81,  # miss
            "Unbalance_V": round(promedio.unbalance_voltage, 2),
            "Unbalance_I": round(promedio.current_unbalance, 2),
            "Frequence": electrical_result_obj.engine_fk.freq_hz,
            "Efficiency": round(promedio.eff_avg, 2),
            "Torque": 620.80,  # miss
            "Load": round(promedio.load_percen_avg, 2),
            "load_kW": 115.7,  # miss
            "Speed": round(electrical_result_obj.engine_fk.speed_rpm, 2)
        }
        return Response(data)
