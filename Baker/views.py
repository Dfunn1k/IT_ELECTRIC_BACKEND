from django.contrib.auth.models import User
from django.contrib.auth.hashers import check_password
from django.core.exceptions import ValidationError
from django.shortcuts import render
from django.utils import timezone
import time
import pandas as pd
from openpyxl import load_workbook
import numpy as np

from concurrent.futures import ThreadPoolExecutor

from rest_framework import status
from rest_framework.authtoken.models import Token
from rest_framework.generics import ListAPIView
from rest_framework.parsers import JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView


from .models import MeasurementER, MeasurementTB, Engine, ElectricalResult, TestER, TestTB, TransientBoot, AverageMeasurement
from .serializers import MeasurementERSerializer, MeasurementTBSerializer, EngineSerializer, ElectricalResultSerializer, TestERSerializer, TestTBSerializer, TransientBootSerializer, UserSerializer

from datetime import datetime

import multiprocessing
import concurrent.futures
from .utils import process_array, data_avarage, create_average, validate_date_format
from openpyxl.utils.exceptions import InvalidFileException, ReadOnlyWorkbookException


class GetUserList(ListAPIView):
    permission_classes = [IsAuthenticated]
    queryset = User.objects.all()
    serializer_class = UserSerializer

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': 'Algo falló al listar los usuarios'},
                            {'exception': e}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class LoginView(APIView):
    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')

        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            return Response({'error': 'Usuario inválido'}, status=status.HTTP_404_NOT_FOUND)

        pwd_valid = check_password(password, user.password)

        if not pwd_valid:
            return Response({'error': 'Contraseña inválida'}, status=status.HTTP_401_UNAUTHORIZED)

        token, _ = Token.objects.get_or_create(user=user)
        return Response({'token': token.key, 'user_id': user.id}, status=status.HTTP_200_OK)


class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            # Elimina el token del usuario
            Token.objects.filter(user=request.user).delete()
            return Response({'detail': 'El token fue eliminado'}, status=status.HTTP_200_OK)
        except Token.DoesNotExist:
            return Response({'error': 'No se encontró un token asociado al usuario'}, status=status.HTTP_400_BAD_REQUEST)


class CreateEngineView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        # Utilizar el JSONParser para procesar los datos en formato JSON
        user = request.user
        parser = JSONParser()
        data = parser.parse(request)

        serializer = EngineSerializer(data=data)
        if Engine.objects.filter(name=data["name"], user=user).exists():
            return Response({'error': 'El usuario ya tiene un motor con ese nombre'}, status=status.HTTP_409_CONFLICT)
        if serializer.is_valid():
            # Obtener o crear el motor según su número nombre
            try:
                engine = Engine.objects.get(
                    name=data["name"], user=request.user)
                response_data = EngineSerializer(engine).data
            except (Engine.DoesNotExist, KeyError):
                # motor = serializer.save(usuario=request.user)
                response_data = serializer.save(user=request.user)

            return Response(response_data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class GetEnginesUserView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, user_pk):
        try:
            engines = Engine.objects.filter(user__pk=user_pk)
            serializer = EngineSerializer(engines, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Engine.DoesNotExist:
            return Response({'error': 'El usuario no tiene motores creados'}, status=status.HTTP_404_NOT_FOUND)


class GetMeasurementsERView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, test_er):
        """Esta función obtiene los datos de las columnas mostradas, re"""
        try:
            measurements = MeasurementER.objects.filter(
                test_electrical_result_fk=test_er)
            data = {
                # 'item': list(mediciones.values_list('item', flat=True)),
                'time': list(measurements.values_list('time', flat=True)),
                'mag_v1': list(map(lambda x: round(x, 2), measurements.values_list('mag_v1', flat=True))),
                'mag_v2': list(map(lambda x: round(x, 2), measurements.values_list('mag_v2', flat=True))),
                'mag_v3': list(map(lambda x: round(x, 2), measurements.values_list('mag_v3', flat=True))),
                'mag_i1': list(map(lambda x: round(x, 2), measurements.values_list('mag_i1', flat=True))),
                'mag_i2': list(map(lambda x: round(x, 2), measurements.values_list('mag_i2', flat=True))),
                'mag_i3': list(map(lambda x: round(x, 2), measurements.values_list('mag_i3', flat=True)))
            }
            return Response(data, status=status.HTTP_200_OK)
        except MeasurementER.DoesNotExist:
            return Response({'error': f'No se encontraron mediciones para el test(pk:{test_er})'}, status=status.HTTP_404_NOT_FOUND)


class GetMeasurementsTBView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, test_tb):
        try:
            measurements = MeasurementTB.objects.filter(
                test_transient_boot_fk=test_tb).order_by('time')
            data = {
                'time': list(measurements.values_list('time', flat=True)),
                'ia': list(measurements.values_list('ia', flat=True)),
                'ib': list(measurements.values_list('ib', flat=True)),
                'ic': list(measurements.values_list('ic', flat=True)),
                'va': list(measurements.values_list('va', flat=True)),
                'vb': list(measurements.values_list('vb', flat=True)),
                'vc': list(measurements.values_list('vc', flat=True))
            }

            data['avg_ia'] = round(sum(data['ia']) / len(data['ia']), 2)
            data['avg_ib'] = round(sum(data['ib']) / len(data['ib']), 2)
            data['avg_ic'] = round(sum(data['ic']) / len(data['ic']), 2)
            data['avg_i'] = round((data['avg_ia'] +
                                   data['avg_ib'] + data['avg_ic']) / 3, 2)
            data['avg_va'] = round(sum(data['va']) / len(data['va']), 2)
            data['avg_vb'] = round(sum(data['vb']) / len(data['vb']), 2)
            data['avg_vc'] = round(sum(data['vc']) / len(data['vc']), 2)
            data['avg_v'] = round((data['avg_va'] +
                                   data['avg_vb'] + data['avg_vc']) / 3, 2)

            return Response(data, status=status.HTTP_200_OK)
        except MeasurementTB.DoesNotExist:
            return Response({'error': f'No se encontraron mediciones para el test(pk: {test_tb})'}, status=status.HTTP_404_NOT_FOUND)


class UploadMeasurementsERView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, pk):
        if "file" not in request.data:
            return Response(
                {'error': 'Ningún archivo proporcionado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.data["file"]

        if not file.name.endswith('.xlsx'):
            return Response(
                {'error': 'El archivo debe tener la extension: .xlsx'},
                status=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE
            )

        try:
            electrical_result_fk = pk
            obj_electrical_result = ElectricalResult.objects.get(
                pk=electrical_result_fk)
        except ElectricalResult.DoesNotExist:
            return Response({'error': f'La instancia de resultado electrico(pk{pk}) no existe '},
                            status=status.HTTP_404_NOT_FOUND)

        try:
            wb = load_workbook(file, read_only=True)
        except (InvalidFileException, ReadOnlyWorkbookException) as e:
            return Response({'error': 'No se pudo cargar el archivo'},
                            {'exception': e},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            sheet_names = wb.sheetnames
            if len(sheet_names) == 0:
                raise IndexError('El archivo no tiene hojas')

            sheet_name = sheet_names[0]
            sheet_name_parts = sheet_name.rsplit('-', 2)
            if len(sheet_name_parts) != 3:
                raise ValueError('Formato de nombre de hoja inválido')

            sheet_name = ":".join(sheet_name_parts)
        except (IndexError, ValueError) as e:
            return Response({'error': e}, status=status.HTTP_400_BAD_REQUEST)

        try:
            test_date_time = validate_date_format(sheet_name)
        except ValueError as e:
            return Response({'error': e}, status=status.HTTP_400_BAD_REQUEST)

        try:
            test_electrical_result = TestER.objects.create(
                electrical_result_fk=obj_electrical_result,
                test_date_time=test_date_time
            )
        except ValidationError as e:
            return Response({'error': e}, status=status.HTTP_409_CONFLICT)

        try:
            df = pd.read_excel(file)
            number_cores = multiprocessing.cpu_count()
        except (PermissionError, IOError):
            return Response({'error': 'File access error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except (pd.errors.ParserError):
            return Response({'error': 'Error parsing Excel file'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        array = df.to_numpy()
        split_array = np.array_split(array, number_cores)
        data_with_test = [(split_array[i], test_electrical_result)
                          for i in range(number_cores)]

        with concurrent.futures.ThreadPoolExecutor(max_workers=number_cores) as executor:
            futures = [executor.submit(process_array, args)
                       for args in data_with_test]
            for future in concurrent.futures.as_completed(futures):
                try:
                    future.result()  # Captura y maneja cualquier excepción lanzada en process_array
                except ValidationError as e:
                    test_electrical_result.delete()
                    return Response({'error': e}, status=status.HTTP_400_BAD_REQUEST)

        try:
            promedio = data_avarage(array, electrical_result_fk)
            create_average(promedio, test_electrical_result)
        except Exception as e:
            test_electrical_result.delete()
            return Response({'error': 'Failed to create averages'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(
            {"message": "Las mediciones E.R han sido creadas exitosamente"},
            status=status.HTTP_201_CREATED,
        )


class UploadMeasurementsTBView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, pk):
        if "file" not in request.data:
            return Response(
                {"error": 'Ningún archivo proporcionado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.data["file"]

        if not file.name.endswith('.xlsx'):
            return Response(
                {'error': 'El archivo debe tener la extension: .xlsx'},
                status=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE
            )

        try:
            transient_boot_fk = pk
            obj_transient_boot = TransientBoot.objects.get(
                pk=transient_boot_fk)
        except TransientBoot.DoesNotExist:
            return Response({'error': f'La instancia de transitorios de arranque(pk:{pk}) no existe'},
                            status=status.HTTP_404_NOT_FOUND)

        try:
            wb = load_workbook(file, read_only=True)
        except (InvalidFileException, ReadOnlyWorkbookException) as e:
            return Response({'error': 'No se pudo cargar el archivo'},
                            {'exception': e},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            sheet_names = wb.sheetnames
            if len(sheet_names) == 0:
                raise IndexError('El archivo no tiene hojas')
            sheet_name = sheet_names[0]
            sheet_name_parts = sheet_name.rsplit('-', 2)
            if len(sheet_name_parts) != 3:
                raise ValueError('Formato de nombre de hoja inválido')

            sheet_name = ":".join(sheet_name_parts)
        except (IndexError, ValueError) as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            test_date_time = validate_date_format(sheet_name)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            test_transient_boot = TestTB.objects.create(
                transient_boot_fk=obj_transient_boot,
                test_date_time=test_date_time
            )
        except ValidationError as e:
            return Response({'error': e}, status=status.HTTP_409_CONFLICT)

        try:
            df = pd.read_excel(file)
            number_cores = multiprocessing.cpu_count()
        except (PermissionError, IOError):
            return Response({'error': 'File access error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except (pd.errors.ParserError):
            return Response({'error': 'Error parsing Excel file'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        array = df.to_numpy()
        split_array = np.array_split(array, number_cores)
        data_with_test = [(split_array[i], test_transient_boot)
                          for i in range(number_cores)]

        with concurrent.futures.ThreadPoolExecutor(max_workers=number_cores) as executor:
            futures = [executor.submit(process_array, args)
                       for args in data_with_test]
            for future in concurrent.futures.as_completed(futures):
                try:
                    future.result()  # Captura y maneja cualquier excepción lanzada en process_array
                except ValidationError as e:
                    test_transient_boot.delete()
                    return Response({'error': e}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {"message": "Las mediciones T.B han sido creadas exitosamente"},
            status=status.HTTP_201_CREATED,
        )


class DeleteEngineView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            engine = Engine.objects.get(pk=pk)
            engine.delete()
            return Response({"message": f"El motor con el id '{pk}' fue correctamente eliminado"}, status=status.HTTP_204_NO_CONTENT)
        except Engine.DoesNotExist:
            return Response({'error': f'El motor {pk} no existe'}, status=status.HTTP_404_NOT_FOUND)


class EditEngineView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        engine = Engine.objects.get(pk=pk)
        serializer = EngineSerializer(engine, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        engine = Engine.objects.get(pk=pk)
        serializer = EngineSerializer(engine, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DeleteTestERView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            test_electrical_result = TestER.objects.get(pk=pk)
            test_electrical_result.delete()
            return Response({"message": f"Las mediciones del test '{pk}' fueron correctamente eliminadas"}, status=status.HTTP_204_NO_CONTENT)
        except TestER.DoesNotExist:
            return Response({'error': f'El test: {pk} no existe'}, status=status.HTTP_404_NOT_FOUND)


class DeleteTestTBView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            test_transient_boot = TestTB.objects.get(pk=pk)
            test_transient_boot.delete()
            return Response({"message": f"Las mediciones del test '{pk}' fueron correctamente eliminadas"}, status=status.HTTP_204_NO_CONTENT)
        except TestTB.DoesNotExist:
            return Response({'error': f'El test: {pk} no existe'}, status=status.HTTP_404_NOT_FOUND)


class AverageView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, test_er):
        try:
            test_er_obj = TestER.objects.get(pk=test_er)
            electrical_result_obj = test_er_obj.electrical_result_fk
            engine_pk = electrical_result_obj.engine_fk.pk
            promedio = AverageMeasurement.objects.get(
                test_electrical_result_fk=test_er)
            tests = TestER.objects.filter(
                electrical_result_fk__engine_fk=engine_pk)
            history = {}
            for test in tests:
                value = AverageMeasurement.objects.get(
                    test_electrical_result_fk=test.pk)
                history[test.pk] = {
                    'value': round(value.value, 2),
                    'fecha': test.test_date_time.strftime("%d/%m/%Y"),
                    'hora': test.test_date_time.strftime("%H:%M:%S"),
                }

            gauge = {
                "minValue": 0,
                "maxValue": 120,
                "value": round(promedio.value, 2)
            }

            data = {
                "popupVoltage": {
                    "values": {
                        "ab": round(promedio.ab, 2),
                        "bc": round(promedio.bc, 2),
                        "ca": round(promedio.ca, 2),
                        "avg": round(promedio.avg, 2),
                        "rated": electrical_result_obj.engine_fk.voltage_rating,
                        "value": round(promedio.value, 2)
                    },
                    "gauge": gauge,
                    "history": history
                },
                "popupUnbalance": {
                    "values": {
                        "ab": round(promedio.ab, 2),
                        "bc": round(promedio.bc, 2),
                        "ca": round(promedio.ca, 2),
                        "avg": round(promedio.avg, 2),
                        "unbalance": round(promedio.unbalance, 2),
                        "nemaDerating": electrical_result_obj.engine_fk.amps_rating,
                    },
                    "gauge": gauge,
                    "history": history
                },
                "popupDistorsion": {
                    "tdhv": {
                        "a": round(promedio.thdv_a, 2),
                        "b": round(promedio.thdv_b, 2),
                        "c": round(promedio.thdv_c, 2),
                        "avg": round(promedio.thdv_avg, 2),
                    },
                    "tdhi": {
                        "a": round(promedio.thdi_a, 2),
                        "b": round(promedio.thdi_b, 2),
                        "c": round(promedio.thdi_c, 2),
                        "avg": round(promedio.thdi_avg, 2)
                    },
                    "history": history
                },
                "popupFullDistorsion": {
                    "tdv": {
                        "a": round(promedio.tdv_a, 2),
                        "b": round(promedio.tdv_b, 2),
                        "c": round(promedio.tdv_c, 2),
                        "avg": round(promedio.tdv_avg, 2)
                    },
                    "tdi": {
                        "a": round(promedio.tdi_a, 2),
                        "b": round(promedio.tdi_b, 2),
                        "c": round(promedio.tdi_c, 2),
                        "avg": round(promedio.tdi_avg, 2)
                    },
                    "gauge": gauge,
                    "history": history
                },
                "popupCurrentLevel": {
                    "values": {
                        "a": round(promedio.current_a, 2),
                        "b": round(promedio.current_b, 2),
                        "c": round(promedio.current_c, 2),
                        "avg": round(promedio.current_avg, 2),
                        "nominalCurrent": round(promedio.current_nominal, 2),
                        "rated": electrical_result_obj.engine_fk.voltage_rating
                    },
                    "gauge": gauge,
                    "history": history
                },
                "popupCurrentUnbalance": {
                    "values": {
                        "a": round(promedio.current_a, 2),
                        "b": round(promedio.current_b, 2),
                        "c": round(promedio.current_c, 2),
                        "avg": round(promedio.current_avg, 2),
                        "currentUnbalance": round(promedio.current_unbalance, 2),
                        "rated": electrical_result_obj.engine_fk.voltage_rating
                    },
                    "gauge": gauge,
                    "history": history
                },
                "popupEfficiency": {
                    "values": {
                        "load": round(promedio.load_percen_avg, 2),
                        "losses": round(promedio.lsskw_avg, 2),
                        "efficiency": round(promedio.eff_avg, 2)
                    },
                    "gauge": gauge,
                    "history": history
                },
                "popupLoad": {
                    "values": {
                        "efficiency": round(promedio.eff_avg, 2),
                        "speed": electrical_result_obj.engine_fk.speed_rpm,
                        "nemaDerating": electrical_result_obj.engine_fk.amps_rating,
                        "load": round(promedio.load_percen_avg, 2)
                    },
                    "gauge": gauge,
                    "history": history
                },
                "popupSpectrum": {
                    "values": {
                        "sideAmplitude": round(promedio.sideband_amplitud_db, 2),
                        "sidebandFreq": round(promedio.sideband_freq_hz, 2),
                        "fundFreq": electrical_result_obj.engine_fk.freq_hz,
                    },
                    "history": history
                },
                "popupEnergia": {
                    "A": {
                        round(promedio.kw_a, 2),
                        round(promedio.kvar_a, 2),
                        round(promedio.kva_a, 2),
                        round(promedio.pf_a, 2),
                        round(promedio.ab, 2),
                        round(promedio.current_a, 2),
                        round(promedio.thdv_a, 2),
                        round(promedio.thdi_a, 2),

                    },
                    "B": {
                        round(promedio.kw_b, 2),
                        round(promedio.kvar_b, 2),
                        round(promedio.kva_b, 2),
                        round(promedio.pf_b, 2),
                        round(promedio.bc, 2),
                        round(promedio.current_b, 2),
                        round(promedio.thdv_b, 2),
                        round(promedio.thdi_b, 2),
                    },
                    "C": {
                        round(promedio.kw_c, 2),
                        round(promedio.kvar_c, 2),
                        round(promedio.kva_c, 2),
                        round(promedio.pf_c, 2),
                        round(promedio.ca, 2),
                        round(promedio.current_c, 2),
                        round(promedio.thdv_c, 2),
                        round(promedio.thdi_c, 2),
                    },
                    "AVg": {
                        round(promedio.kw_avg, 2),
                        round(promedio.kvar_avg, 2),
                        round(promedio.kva_avg, 2),
                        round(promedio.pf_a, 2),
                        round(promedio.avg, 2),
                        round(promedio.current_avg, 2),
                        round(promedio.thdv_avg, 2),
                        round(promedio.thdi_avg, 2),
                    },
                    "unbalance_voltage": round(promedio.unbalance_voltage, 2),
                    "unabalance_current": round(promedio.unbalance_current, 2),
                    "frecuence": round(promedio.sideband_freq_hz, 2)
                },
                "SymmetricalComponents": {
                    "title": [
                        "Vab",
                        "Vbc",
                        "Vca",
                        "IA",
                        "IB",
                        "IC"
                    ],
                    "Amplitud": [
                        round(promedio.ab, 2),
                        round(promedio.bc, 2),
                        round(promedio.ca, 2),
                        round(promedio.current_a, 2),
                        round(promedio.current_b, 2),
                        round(promedio.current_c, 2),
                    ],
                    "Fase": [
                        round(promedio.vab_fase, 2),
                        round(promedio.vbc_fase, 2),
                        round(promedio.vca_fase, 2),
                        round(promedio.ia_fase, 2),
                        round(promedio.ib_fase, 2),
                        round(promedio.ic_fase, 2),
                    ],
                    "Desbalance": [
                        round(promedio.unbalance_voltage, 2),
                        round(promedio.unbalance_current, 2)
                    ],
                    "Amplitud_2": [
                        round(promedio.va1_amplitud, 2),
                        round(promedio.va2_amplitud, 2),
                        round(promedio.ia1_amplitud, 2),
                        round(promedio.ia2_amplitud, 2),
                    ],
                    "Fase_2": [
                        round(promedio.va1_fase, 2),
                        round(promedio.va2_fase, 2),
                        round(promedio.ia1_fase, 2),
                        round(promedio.ia2_fase, 2)
                    ]
                }
            }

            return Response(data, status=status.HTTP_200_OK)
        except TestER.DoesNotExist:
            return Response({'error': f'El test {test_er} no existe'}, status=status.HTTP_404_NOT_FOUND)


class MainView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, test_er):
        try:
            test_er_obj = TestER.objects.get(pk=test_er)
            electrical_result_obj = test_er_obj.electrical_result_fk
            engine_pk = electrical_result_obj.engine_fk.pk
            promedio = AverageMeasurement.objects.get(
                test_electrical_result_fk=test_er)
            tests = TestER.objects.filter(
                electrical_result_fk__engine_fk=engine_pk)

            data = {
                "engine": {
                    "name": electrical_result_obj.engine_fk.name,
                    "voltage": f"{electrical_result_obj.engine_fk.voltage_rating}",
                    "current": f"{electrical_result_obj.engine_fk.amps_rating}",
                    "kW": f"{electrical_result_obj.engine_fk.power_out_kw}",
                    "RPM": f"{electrical_result_obj.engine_fk.speed_rpm}",
                },
                "Voltage": round(promedio.avg, 2),
                "Current": round(promedio.current_avg, 2),
                "FP": round(promedio.pf_avg, 2),
                "Unbalance_V": round(promedio.unbalance_voltage, 2),
                "Unbalance_I": round(promedio.current_unbalance, 2),
                "Frequence": electrical_result_obj.engine_fk.freq_hz,
                "Efficiency": round(promedio.eff_avg, 2),
                "Torque": round(promedio.torque, 2),
                "Load": round(promedio.load_percen_avg, 2),
                "load_kW": 115.7,  # miss
                "Speed": round(electrical_result_obj.engine_fk.speed_rpm, 2)
            }
            return Response(data, status=status.HTTP_200_OK)
        except TestER.DoesNotExist:
            return Response({'error': f'El test {test_er} no existe'}, status=status.HTTP_404_NOT_FOUND)
